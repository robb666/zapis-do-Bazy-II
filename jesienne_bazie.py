import requests
import win32com.client
from win32com.client import Dispatch
import time
from openpyxl import load_workbook
import datetime
import re
from creds import key
from icecream import ic


start_time = time.time()


class PyxlExcel:
    def __init__(self, filename=None, workbook=None):
        self.wb = load_workbook(filename)
        self.ws = self.wb[workbook]

    def get_cell(self, col_row):
        return self.ws[col_row].value

    def close(self):
        return self.wb.close()


class Win32comExcel:
    def __init__(self, filename=None, workbook=None, sheet=None, visible=True):
        try:
            self.ExcelApp = win32com.client.GetActiveObject('Excel.Application')
            self.wb = self.ExcelApp.Workbooks(workbook)
            self.ws = self.wb.Worksheets(sheet)
        except:
            self.ExcelApp = win32com.client.gencache.EnsureDispatch('Excel.Application')
            self.wb = self.ExcelApp.Workbooks.OpenXML(filename)
            self.ws = self.wb.Worksheets(sheet)
        self.ExcelApp.Visible = visible

    def get_next_row(self, col):
        return self.ws.Cells(self.ws.Rows.Count, col).End(-4162).Row + 1

    def get_last_cell(self, col):
        return self.ws.Cells(self.ws.Rows.Count, col).End(-4162)

    def date_formatter(self, date):
        if date not in ('', None, 'None'):
            day, month, year = date.split('.')
            return datetime.datetime(int(year), int(month), int(day)).strftime('%Y-%m-%d')
        return ''

    def row_range_input(self, data):
        row_to_write = self.get_next_row(col=30)
        self.ws.Range(self.ws.Cells(row_to_write, 1), self.ws.Cells(row_to_write, len(data))).Value = data


class ValidatedAPIRequester:
    def __init__(self, base_url, headers):
        self.base_url = base_url
        self.headers = headers
        self.endpoints = {}  # Store endpoints

    def add_endpoint(self, name, endpoint):
        self.endpoints[name] = endpoint

    def __getitem__(self, key):
        return self.endpoints.get(key)

    def post(self, endpoint, data):
        return requests.post(self.base_url + endpoint, headers=self.headers, json=data).json()

    def regon_checksum(self, r: str):
        if len(r) == 9:
            regon = list(str(r))
            suma = (int(regon[0]) * 8 + int(regon[1]) * 9 + int(regon[2]) * 2 + int(regon[3]) * 3 + int(regon[4]) * 4 +
                    int(regon[5]) * 5 + int(regon[6]) * 6 + int(regon[7]) * 7) % 11
            if suma == int(regon[-1]) or suma == 10 and int(regon[-1]) == 0:
                return True
            else:
                return False
        else:
            return False

    def pesel_checksum(self, p):
        if len(p) == 11:
            l = int(p[10])
            suma = 1 * int(p[0]) + 3 * int(p[1]) + 7 * int(p[2]) + 9 * int(p[3]) + 1 * int(p[4]) + 3 * int(p[5]) + \
                   7 * int(p[6]) + 9 * int(p[7]) + 1 * int(p[8]) + 3 * int(p[9]) + 1 * int(p[10])
            lm = suma % 10
            kontrola = 10 - lm
            if (kontrola == 10 or l == kontrola) and p[2:4] != '00':
                return True
            else:
                return False
        else:
            return False

    def insurer(self, tow):
        insurers = {'Allianz': 'ALL', 'AXA': 'AXA', 'Balcia': 'BAL', 'Compensa': 'COM', 'Euroins': 'EIN',
                    'I10001237': 'PZU', 'pzu': 'PZU', 'Generali': 'GEN', 'HDI': 'HDI',
                    'Ergo Hestia': 'HES', 'Ergohestialite': 'HES', 'INTER': 'INT', 'LINK 4': 'LIN',
                    'mtu': 'MTU', 'Proama': 'PRO', 'InterRisk': 'RIS', 'tuwtuw': 'TUW',
                    'tuz': 'TUZ', 'Uniqa': 'UNI', 'Warta': 'WAR', 'Wiener': 'WIE', 'Gothaer': 'WIE',
                    'You Can Drive': 'YCD', 'Trasti': 'TRA', 'Wefox': 'WEF'}
        for insurer in insurers:
            if re.search(tow, insurer, re.I):
                return insurers[insurer]
        else:
            return ''

    # TODO rozdzielić markę i model (plik w starej wersji)
    def insurance_type(self, rodzaj):
        moto = 'Ubezpieczenie komunikacyjne motor'
        if re.search(rodzaj, moto, re.I):
            return 'kom'
        else:
            return ''


pyxl = PyxlExcel(
    filename='M:/Agent baza/Login_Hasło.xlsm',
    workbook='Aplikacje',
)

ExcelApp = Win32comExcel(
    filename='C:\\Users\\PipBoy3000\\Desktop\\2014 BAZA MAGRO.xlsm',
    workbook='2014 BAZA MAGRO.xlsm',
    sheet='BAZA 2014',
)

in_l = pyxl.get_cell('F21')
in_h = pyxl.get_cell('G21')
pyxl.close()

from_date = ExcelApp.get_last_cell(col=30)


str_conv = str(from_date)[:10].replace('.', '-')
year, month, day = str_conv.split('-')
timestamp_from = datetime.datetime(int(year), int(month), int(day)).strftime('%d.%m.%Y')
timestamp_to = datetime.date.today().strftime('%d.%m.%Y')


headers = {
    "Authorization": f"Bearer {key}",
    "Content-Type": "application/json",
}

policy_list_payload = {
    "username": in_l,
    "password": in_h,
    "ajax_url": "/api/policy/list",
    "output": "json",
    "timestamp_from": "15.11.2023", #,  # timestamp_from
    "timestamp_to": "15.12.2023", #timestamp_to,
}

api_requester = ValidatedAPIRequester(
        base_url='https://magro2-api.insly.pl/api/',
        headers=headers,
)

api_requester.add_endpoint('policies list', 'policy/list')
policies_list = api_requester.post(api_requester['policies list'],
                                   data=policy_list_payload)
print(policies_list)
api_requester.add_endpoint('get policy', 'policy/getpolicy')
print(policies_list['policies'])

for policy in policies_list['policies']:
    policy_oid = policy['policy_oid']

    payload = {
        "ajax_url": "/api/policy/getpolicy",
        "output": "json",
        "policy_oid": policy_oid,
        "return_objects": "1"
    }

    r = api_requester.post(api_requester['get policy'], data=payload)
    ic(r)

    pesel = api_requester.pesel_checksum(r['customer_idcode'])
    regon = api_requester.regon_checksum(r['customer_idcode'])
    nazwa_firmy = r['customer_name'] if regon else ''
    nazwisko = r['customer_name'].split()[-1] if nazwa_firmy == '' else ''
    imie = r['customer_name'].split()[0] if nazwa_firmy == '' else ''
    p_lub_r = r['customer_idcode'] if pesel else r['customer_idcode'] if regon else ''
    ulica = r['address'][0]['customer_address_street']
    kod_poczt = r['address'][0]['customer_address_zip']
    miasto = r['address'][0]['customer_address_city']
    tel = r['customer_mobile'] if r['customer_mobile'] != '' else r['customer_phone']
    tel = tel.lstrip('+48')
    email = r['customer_email']

    marka = r.get('objects')[0].get('vehicle_make', '') if len(r['objects']) > 0 else ''
    model = r.get('objects', '')[0].get('vehicle_model', '') if len(r['objects']) > 0 else ''
    nr_rej = ''
    rok = ''
    if len(r['objects']) > 0:
        nr_rej = r.get('objects')[0].get('vehicle_registration_number', '') \
            if r.get('objects')[0].get('vehicle_registration_number', '') != '' \
            else r.get('objects')[0].get('vehicle_licenseplate', '')

        rok = r.get('objects')[0].get('vehicle_first_registration_date', '')[:4] \
            if r.get('objects')[0].get('vehicle_first_registration_date', '') != '' \
            else r.get('objects')[0].get('vehicle_registered', '')[:4]

    data_pocz = ExcelApp.date_formatter(r.get('policy_date_start', ''))
    data_konca = ExcelApp.date_formatter(r.get('policy_date_end', ''))
    tow_ub = api_requester.insurer(r.get('policy_insurer', ''))
    rodzaj = api_requester.insurance_type(r.get('policy_product_info', '')[0].get('policy_product_displayname', ''))
    nr_polisy = r['policy_no']
    przypis = r['policy_payment_sum']
    ter_platnosci = ExcelApp.date_formatter(r.get('payment', '')[0].get('policy_installment_date_due', ''))
    f_platnosci = 'P' if r.get('policy_first_installment_payment_method') == 3 else ''
    ilosc_rat = r.get('policy_installments', '')

    I_rata = r.get('payment')[0].get('policy_installment_sum_real', '')
    nr_raty = '1'

    print(nazwa_firmy)
    print(nazwisko)
    print(imie)
    print(p_lub_r)
    print(ulica)
    print(kod_poczt)
    print(miasto)
    print(tel)
    print(email)
    print(marka)
    print(model)
    print(nr_rej)
    print(rok)
    print(data_pocz)
    print(data_konca)
    print(tow_ub)
    print(rodzaj)

    print('-------------')

    data = [
        '', '', '', '', '', '',
        'Robert', '', '',
        'Grzelak',
        nazwa_firmy,
        nazwisko,
        imie,
        pesel_lub_regon := 'p' + p_lub_r if len(p_lub_r) == 11 else 'r' + p_lub_r if len(p_lub_r) == 9 else '', '',
        ulica,
        kod_poczt,
        miasto,
        tel,
        email := email.lower() if email else '', '', '',
        marka if nr_rej != '' else kod_poczt,
        model if nr_rej != '' else miasto,
        nr_rej if nr_rej != '' else ulica,
        rok, '', '', '',
        datetime.date.today().strftime('%Y-%m-%d'),
        data_pocz,
        data_konca, '', '', '',
        'SPÓŁKA',
        tow_ub,
        tow_ub,
        'kom' if nr_rej != '' else '',
        nr_polisy, '', '', '', '', '', '', '',
        ### logika rat ###
        przypis,
        ter_platnosci,
        I_rata if I_rata else przypis,
        f_platnosci,
        ilosc_rat,
        nr_raty,
        ter_platnosci, '', '', '',
        'api', '',
        tow_ub
    ]  # data for the row



    # TODO zaimplementować to w Win32comExcel i zrobić wszystkie raty..nawet 12
    ExcelApp.row_range_input(data)

    """  RATY  """

    II_rata, III_rata, IV_rata = '', '', ''

    row_to_write = ExcelApp.get_next_row(col=30)

    if len(r.get('payment')) > 1:
        II_rata = r.get('payment')[1].get('policy_installment_sum_real', '')
        II_ter_platnosci = ExcelApp.date_formatter(r.get('payment', '')[1].get('policy_installment_date_due', ''))
        nr_raty = '2'
        II_rata_data = data[:-13] + ['', II_ter_platnosci, II_rata, f_platnosci, ilosc_rat, nr_raty, '', '', '', '',
                                     'api', '', tow_ub]
        ExcelApp.row_range_input(II_rata_data)

    if len(r.get('payment')) > 2:
        III_rata = r.get('payment')[2].get('policy_installment_sum_real', '')
        III_ter_platnosci = ExcelApp.date_formatter(r.get('payment', '')[2].get('policy_installment_date_due', ''))
        nr_raty = '3'
        III_rata_data = data[:-13] + ['', III_ter_platnosci, III_rata, f_platnosci, ilosc_rat, nr_raty, '', '', '', '',
                                     'api', '', tow_ub]
        ExcelApp.row_range_input(III_rata_data)

    if len(r.get('payment')) > 3:
        IV_rata = r.get('payment')[3].get('policy_installment_sum_real', '')
        IV_ter_platnosci = ExcelApp.date_formatter(r.get('payment', '')[3].get('policy_installment_date_due', ''))
        nr_raty = '4'
        IV_rata_data = data[:-13] + ['', IV_ter_platnosci, IV_rata, f_platnosci, ilosc_rat, nr_raty, '', '', '', '',
                                     'api', '', tow_ub]
        ExcelApp.row_range_input(IV_rata_data)

    if len(r.get('payment')) > 4:
        V_rata = r.get('payment')[4].get('policy_installment_sum_real', '')
        V_ter_platnosci = ExcelApp.date_formatter(r.get('payment', '')[4].get('policy_installment_date_due', ''))
        nr_raty = '4'
        V_rata_data = data[:-13] + ['', V_ter_platnosci, V_rata, f_platnosci, ilosc_rat, nr_raty, '', '', '', '',
                                     'api', '', tow_ub]
        ExcelApp.row_range_input(V_rata_data)

    if len(r.get('payment')) > 5:
        VI_rata = r.get('payment')[5].get('policy_installment_sum_real', '')
        VI_ter_platnosci = ExcelApp.date_formatter(r.get('payment', '')[5].get('policy_installment_date_due', ''))
        nr_raty = '4'
        VI_rata_data = data[:-13] + ['', VI_ter_platnosci, VI_rata, f_platnosci, ilosc_rat, nr_raty, '', '', '', '',
                                     'api', '', tow_ub]
        ExcelApp.row_range_input(VI_rata_data)

    if len(r.get('payment')) > 6:
        VII_rata = r.get('payment')[6].get('policy_installment_sum_real', '')
        VII_ter_platnosci = ExcelApp.date_formatter(r.get('payment', '')[6].get('policy_installment_date_due', ''))
        nr_raty = '4'
        VII_rata_data = data[:-13] + ['', VII_ter_platnosci, VII_rata, f_platnosci, ilosc_rat, nr_raty, '', '', '', '',
                                     'api', '', tow_ub]
        ExcelApp.row_range_input(VII_rata_data)

    if len(r.get('payment')) > 7:
        VIII_rata = r.get('payment')[7].get('policy_installment_sum_real', '')
        VIII_ter_platnosci = ExcelApp.date_formatter(r.get('payment', '')[7].get('policy_installment_date_due', ''))
        nr_raty = '4'
        VIII_rata_data = data[:-13] + ['', VIII_ter_platnosci, VIII_rata, f_platnosci, ilosc_rat, nr_raty, '', '', '', '',
                                     'api', '', tow_ub]
        ExcelApp.row_range_input(VIII_rata_data)

    if len(r.get('payment')) > 8:
        IX_rata = r.get('payment')[8].get('policy_installment_sum_real', '')
        IX_ter_platnosci = ExcelApp.date_formatter(r.get('payment', '')[8].get('policy_installment_date_due', ''))
        nr_raty = '4'
        IX_rata_data = data[:-13] + ['', IX_ter_platnosci, IX_rata, f_platnosci, ilosc_rat, nr_raty, '', '', '', '',
                                     'api', '', tow_ub]
        ExcelApp.row_range_input(IX_rata_data)

    if len(r.get('payment')) > 9:
        X_rata = r.get('payment')[9].get('policy_installment_sum_real', '')
        X_ter_platnosci = ExcelApp.date_formatter(r.get('payment', '')[9].get('policy_installment_date_due', ''))
        nr_raty = '4'
        X_rata_data = data[:-13] + ['', X_ter_platnosci, X_rata, f_platnosci, ilosc_rat, nr_raty, '', '', '', '',
                                     'api', '', tow_ub]
        ExcelApp.row_range_input(X_rata_data)

    if len(r.get('payment')) > 10:
        XI_rata = r.get('payment')[10].get('policy_installment_sum_real', '')
        XI_ter_platnosci = ExcelApp.date_formatter(r.get('payment', '')[10].get('policy_installment_date_due', ''))
        nr_raty = '4'
        XI_rata_data = data[:-13] + ['', XI_ter_platnosci, XI_rata, f_platnosci, ilosc_rat, nr_raty, '', '', '', '',
                                     'api', '', tow_ub]
        ExcelApp.row_range_input(XI_rata_data)

    if len(r.get('payment')) > 12:
        XII_rata = r.get('payment')[12].get('policy_installment_sum_real', '')
        XII_ter_platnosci = ExcelApp.date_formatter(r.get('payment', '')[12].get('policy_installment_date_due', ''))
        nr_raty = '4'
        XII_rata_data = data[:-13] + ['', XII_ter_platnosci, XII_rata, f_platnosci, ilosc_rat, nr_raty, '', '', '', '',
                                     'api', '', tow_ub]
        ExcelApp.row_range_input(XII_rata_data)





    # if rata_I:
    #     ExcelApp.Cells(row_to_write, 50).Value = rata_I
    # else:
    #     ExcelApp.Cells(row_to_write, 50).Value = przypis
    # ExcelApp.Cells(row_to_write, 51).Value = f_platnosci
    # ExcelApp.Cells(row_to_write, 52).Value = ilosc_rat
    # ExcelApp.Cells(row_to_write, 53).Value = nr_raty
    # data_inkasa = ExcelApp.Cells(row_to_write, 54).Value = ter_platnosci
    # if rata_I:
    #     ExcelApp.Cells(row_to_write, 55).Value = rata_I
    # else:
    #     ExcelApp.Cells(row_to_write, 55).Value = przypis
    # ExcelApp.Cells(row_to_write, 58).Value = 'aut'
    # ExcelApp.Cells(row_to_write, 60).Value = tow_ub_tor
    #
    # if rata_II:
    #     ws.Range(f'A{row_to_write}:BH{row_to_write}').Copy()
    #     ws.Range(f'A{row_to_write + 1}').PasteSpecial()
    #
    #     ExcelApp.Cells(row_to_write + 1, 48).Value = ''
    #     ExcelApp.Cells(row_to_write + 1, 49).Value = termin_II
    #     ExcelApp.Cells(row_to_write + 1, 50).Value = rata_II
    #     ExcelApp.Cells(row_to_write + 1, 53).Value = 2
    #     data_inkasa = ExcelApp.Cells(row_to_write + 1, 54).Value = ''
    #     ExcelApp.Cells(row_to_write + 1, 55).Value = ''
    #
    #     if rata_III:
    #         ws.Range(f'A{row_to_write + 1}:BH{row_to_write + 1}').Copy()
    #         ws.Range(f'A{row_to_write + 2}').PasteSpecial()
    #
    #         ExcelApp.Cells(row_to_write + 2, 49).Value = termin_III
    #         ExcelApp.Cells(row_to_write + 2, 50).Value = rata_III
    #         ExcelApp.Cells(row_to_write + 2, 53).Value = 3
    #
    #         if rata_IV:
    #             ws.Range(f'A{row_to_write + 2}:BH{row_to_write + 2}').Copy()
    #             ws.Range(f'A{row_to_write + 3}').PasteSpecial()
    #
    #             ExcelApp.Cells(row_to_write + 3, 49).Value = termin_IV
    #             ExcelApp.Cells(row_to_write + 3, 50).Value = rata_IV
    #             ExcelApp.Cells(row_to_write + 3, 53).Value = 4



    # row_to_write += 1



# """Opcje zapisania"""
#
# # Exec
# ExcelApp.DisplayAlerts = False
#
# ## Bez zapisania
# # wb.SaveAs("M:\\Agent baza\\2014 BAZA MAGRO.xlsm")
#
# # Testy
# # wb.SaveAs(path + "\\2014 BAZA MAGRO.xlsm")
#
# """Zamknięcie narazie wyłączone..."""
# # wb.Close()
# ExcelApp.DisplayAlerts = True
#
# # ExcelApp.Application.Quit()
#
# end_time = time.time() - start_time
# print('Czas wykonania: {:.2f} sekund'.format(end_time))