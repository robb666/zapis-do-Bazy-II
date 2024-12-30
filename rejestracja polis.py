import requests
import win32com.client
import win32com.client as win32
# from win32com.client import Dispatch
import time
from openpyxl import load_workbook
import datetime
import re
from creds import key
# from icecream import ic


start_time = time.time()


class PyxlExcel:
    def __init__(self, filename=None, workbook=None):
        self.wb = load_workbook(filename)
        self.ws = self.wb[workbook]

    def get_cell(self, col_row):
        return self.ws[col_row].value

    def close(self):
        return self.wb.close()


class Win32comExcel:

    def __init__(self, filename=None, workbook=None, sheet=None, visible=True):
        self.ExcelApp = win32com.client.gencache.EnsureDispatch('Excel.Application')
        try:
            self.ExcelApp = win32com.client.GetActiveObject('Excel.Application')
            self.wb = self.ExcelApp.Workbooks(workbook)
            self.ws = self.wb.Worksheets(sheet)
        except:
            self.wb = self.ExcelApp.Workbooks.OpenXML(filename)
            self.ws = self.wb.Worksheets(sheet)
        self.ExcelApp.Visible = visible

    def get_last_cell_value(self, col):
        return self.ws.Cells(self.ws.Rows.Count, col).End(-4162)

    def _get_next_row(self, col):
        return self.ws.Cells(self.ws.Rows.Count, col).End(-4162).Row + 1

    def _find_last_row_by_value(self, col):
        last_row = self.ws.Cells(self.ws.Rows.Count, col).End(-4162).Row  # .End(win32com.client.constants.xlUp).Row
        for row in range(last_row, 0, -1):
            if self.ws.Cells(row, col).Value in ('MAGRO', 'Wawrzyniak', 'Wołowski', 'Skrzypek', 'Nowakowski', 'Filipiak'):
                return row + 1

    def date_formatter(self, date):
        if date not in ('', None, 'None'):
            day, month, year = date.split('.')
            return datetime.datetime(int(year), int(month), int(day)).strftime('%Y-%m-%d')
        return ''

    def row_range_input(self, data, value=False):
        row_to_write = self._get_next_row(col=30)
        if value:
            row_to_write = self._find_last_row_by_value(col=7)
            self.ws.Rows(row_to_write).Copy()
            self.ws.Rows(row_to_write).Insert(Shift=win32.constants.xlDown)
        data[22] = f'=AF{row_to_write}-AE{row_to_write}+1'
        data[26] = f'=IF(OR(AC{row_to_write}="anulowana",AF{row_to_write}="",AT{row_to_write}=""),"",(IF(AF{row_to_write}+10<NOW(),"po_10",AF{row_to_write})))'
        data[-4] = f'=AX{row_to_write}-BC{row_to_write}'
        self.ws.Range(self.ws.Cells(row_to_write, 'G'), self.ws.Cells(row_to_write, 'BH')).Value = data


class ValidatedAPIRequester:
    def __init__(self, base_url, headers):
        self.base_url = base_url
        self.headers = headers
        self.endpoints = {}  # Store endpoints

    def add_endpoint(self, name, endpoint):
        self.endpoints[name] = endpoint

    def __getitem__(self, key):
        return self.endpoints.get(key)

    def post(self, endpoint, data):
        return requests.post(self.base_url + endpoint, headers=self.headers, json=data).json()

    def ofwce(self, ofwca_id):
        ofwce_ids = {
            10001198: {
                'Name': 'Robert',
                'Surname': 'Grzelak'
            },
            10001211: {
                'Name': 'MAGRO',
                'Surname': 'Grzelak'
            },
            'x': {
                'Name': 'Agnieszka',
                'Surname': 'Wawrzyniak'
            },
        }

        if ofwca_id in ofwce_ids:
            return ofwce_ids[ofwca_id]
        else:
            ''

    def regon_checksum(self, r: str):
        if len(r) == 9:
            regon = list(str(r))
            suma = (int(regon[0]) * 8 + int(regon[1]) * 9 + int(regon[2]) * 2 + int(regon[3]) * 3 + int(regon[4]) * 4 +
                    int(regon[5]) * 5 + int(regon[6]) * 6 + int(regon[7]) * 7) % 11
            if suma == int(regon[-1]) or suma == 10 and int(regon[-1]) == 0:
                return True
            else:
                return False
        else:
            return False

    def pesel_checksum(self, p):
        if len(p) == 11:
            l = int(p[10])
            suma = 1 * int(p[0]) + 3 * int(p[1]) + 7 * int(p[2]) + 9 * int(p[3]) + 1 * int(p[4]) + 3 * int(p[5]) + \
                   7 * int(p[6]) + 9 * int(p[7]) + 1 * int(p[8]) + 3 * int(p[9]) + 1 * int(p[10])
            lm = suma % 10
            kontrola = 10 - lm
            if (kontrola == 10 or l == kontrola) and p[2:4] != '00':
                return True
            else:
                return False
        else:
            return False

    def insurer(self, tow):
        insurers = {'Allianz': 'ALL', 'AXA': 'AXA', 'Balcia': 'BAL', 'Compensa': 'COM', 'Euroins': 'EIN',
                    'I10001237': 'PZU', 'pzu': 'PZU', 'Generali': 'GEN', 'HDI': 'HDI',
                    'Ergo Hestia': 'HES', 'Ergohestialite': 'HES', 'INTER': 'INT', 'LINK 4': 'LIN',
                    'mtu': 'MTU', 'Proama': 'PRO', 'InterRisk': 'RIS', 'tuwtuw': 'TUW',
                    'tuz': 'TUZ', 'Uniqa': 'UNI', 'Warta': 'WAR', 'Wiener': 'WIE', 'Gothaer': 'WIE',
                    'You Can Drive': 'YCD', 'Trasti': 'TRA', 'Wefox': 'WEF'}
        for insurer in insurers:
            if re.search(tow, insurer, re.I):
                return insurers[insurer]
        else:
            return ''

    def insurance_type(self, rodzaj):
        moto = ('Ubezpieczenie komunikacyjne', 'ubezpieczenie komunikacyjne', 'Motor (w/o calculation)')
        if rodzaj in moto:
            return 'kom'
        else:
            return ''

    def car_make_model(self, policy_description):
        if policy_description:
            with open('M:\\Agent baza\\marki.txt') as content:
                makes = content.read().split('\n')
                for make in makes:
                    if re.search(make, policy_description, re.I):
                        if model := re.search(rf'{make},?\s([\w\s\d-]+)', policy_description):
                            model = model.group(1)
                        return make, model
                return '', ''


pyxl = PyxlExcel(
    filename='M:/Agent baza/Login_Hasło.xlsm',
    workbook='Aplikacje',
)

ExcelApp = Win32comExcel(
    # filename='C:\\Users\\PipBoy3000\\Desktop\\2014 BAZA MAGRO.xlsm',
    filename='M:\\Agent baza\\2014 BAZA MAGRO.xlsm',
    workbook='2014 BAZA MAGRO.xlsm',
    sheet='BAZA 2014',
)

in_l = pyxl.get_cell('F21')
in_h = pyxl.get_cell('G21')
pyxl.close()


from_date = ExcelApp.get_last_cell_value(col=30)
str_conv = str(from_date)[:10].replace('.', '-')
year, month, day = str_conv.split('-')
timestamp_from = datetime.datetime(int(year), int(month), int(day))

from_next_day = (timestamp_from + datetime.timedelta(days=1)).strftime('%d.%m.%Y')
timestamp_to = (datetime.date.today() + datetime.timedelta(days=1)).strftime('%d.%m.%Y')


headers = {
    "Authorization": f"Bearer {key}",
    "Content-Type": "application/json",
}

policy_list_payload = {
    "username": in_l,
    "password": in_h,
    "ajax_url": "/api/policy/list",
    "output": "json",
    "timestamp_from": from_next_day,
    "timestamp_to": timestamp_to,
}

api_requester = ValidatedAPIRequester(
        base_url='https://magro2-api.insly.pl/api/',
        headers=headers,
)

api_requester.add_endpoint('policies list', 'policy/list')
policies_list = api_requester.post(api_requester['policies list'],
                                   data=policy_list_payload)

api_requester.add_endpoint('get policy', 'policy/getpolicy')

print('\nPobieranie danych polis z CRM przez API')

for policy in policies_list['policies']:
    policy_oid = policy['policy_oid']

    payload = {
        "ajax_url": "/api/policy/getpolicy",
        "output": "json",
        "policy_oid": policy_oid,
        "return_objects": "1"
    }

    r = api_requester.post(api_requester['get policy'], data=payload)
    # ic(r)

    ofwca = api_requester.ofwce(r.get('sales_broker_person_oid', ''))
    #     r.get('broker_person_oid', '')
    #     if r.get('broker_person_oid', '') != ''
    #     else r.get('sales_broker_person_oid', '')
    # )

    pesel = api_requester.pesel_checksum(r['customer_idcode'])
    regon = api_requester.regon_checksum(r['customer_idcode'])
    nazwa_firmy = r['customer_name'] if regon else ''
    nazwisko = r['customer_name'].split()[-1] if nazwa_firmy == '' else ''
    imie = r['customer_name'].split()[0] if nazwa_firmy == '' else ''
    p_lub_r = r['customer_idcode'] if pesel else r['customer_idcode'] if regon else ''
    pesel_lub_regon = 'p' + p_lub_r if len(p_lub_r) == 11 else 'r' + p_lub_r if len(p_lub_r) == 9 else ''
    ulica = r.get('address', '')[0]['customer_address_street'] if r.get('address') else ''
    nr_ulicy = r.get('address', '')[0]['customer_address_house'] if r.get('address') else ''
    nr_mie = r.get('address', '')[0].get('customer_address_apt', '') if r.get('address') else ''
    adres = f'{ulica} {nr_ulicy}'
    if nr_mie not in ('None', None):
        adres = f'{ulica} {nr_ulicy} m {nr_mie}'
    kod_poczt = r.get('address', '')[0]['customer_address_zip'] if r.get('address') else ''
    miasto =r.get('address', '')[0]['customer_address_city'] if r.get('address') else ''
    tel = r['customer_mobile'] if r['customer_mobile'] != '' else r['customer_phone']
    tel = tel.lstrip('+48')
    email = r['customer_email']

    marka = r.get('objects')[0].get('vehicle_make', '') if len(r['objects']) > 0 else ''
    if marka == '':
        marka = api_requester.car_make_model(r.get('policy_description', None))[0]
    model = r.get('objects', '')[0].get('vehicle_model', '') if len(r['objects']) > 0 else ''
    if model == '':
        model = api_requester.car_make_model(r.get('policy_description', None))[1]
    nr_rej = ''
    rok = ''
    if len(r['objects']) > 0:
        nr_rej = r.get('objects')[0].get('vehicle_registration_number', '') \
            if r.get('objects')[0].get('vehicle_registration_number', '') != '' \
            else r.get('objects')[0].get('vehicle_licenseplate', '')

        rok = r.get('objects')[0].get('vehicle_first_registration_date', '')[:4] \
            if r.get('objects')[0].get('vehicle_first_registration_date', '') != '' \
            else r.get('objects')[0].get('vehicle_registered', '')[:4]

    data_pocz = ExcelApp.date_formatter(r.get('policy_date_start', ''))
    data_konca = ExcelApp.date_formatter(r.get('policy_date_end', ''))
    tow_ub = api_requester.insurer(r.get('policy_insurer', ''))
    rodzaj = api_requester.insurance_type(
        r.get('policy_product_info', '')[0].get('policy_product_displayname', '')
        if r.get('policy_product_info', '')[0].get('policy_product_displayname', '') != ''
        else r.get('policy_type', ''))
    nr_polisy = r['policy_no']
    przypis = r['policy_payment_sum']
    ter_platnosci = ExcelApp.date_formatter(r.get('payment', '')[0].get('policy_installment_date_due', ''))
    f_platnosci = 'P' if r.get('policy_first_installment_payment_method') == 3 else \
        'G' if r.get('policy_first_installment_payment_method') == 1 \
            else ''
    ilosc_rat = r.get('policy_installments', '')

    I_rata = r.get('payment')[0].get('policy_installment_sum_real', '')
    nr_raty = '1'

    print('-------------')
    print('  Nazwa firmy: ', nazwa_firmy)
    print('     Nazwisko: ', nazwisko)
    print('         Imię: ', imie)
    print('Pesel / Regon: ', pesel_lub_regon)
    print('      Telefon: ', tel)
    print('        Email: ', email)

    data = [
        # od 7 kolumny 'G'
        ofwca['Name'], '', '',
        ofwca['Surname'],
        nazwa_firmy,
        nazwisko,
        imie,
        pesel_lub_regon, '',
        adres,
        kod_poczt,
        miasto,
        tel,
        email := email.lower() if email else None, None, None,
        marka if nr_rej != '' else kod_poczt,
        model if nr_rej != '' else miasto,
        nr_rej if nr_rej != '' else ulica,
        rok, '', '', None,
        datetime.date.today().strftime('%Y-%m-%d'),
        data_pocz,
        data_konca, None, '', '',
        'SPÓŁKA',
        tow_ub,
        tow_ub,
        rodzaj,
        nr_polisy, '', '', '', '', '', '', '',
        ### logika rat ###
        przypis,
        ter_platnosci,
        I_rata if I_rata else przypis,
        f_platnosci,
        ilosc_rat,
        nr_raty,
        ter_platnosci,
        zainkasowana_rata := I_rata if I_rata else przypis, None, None,
        'api', '',
        tow_ub
    ]  # data for the row

    if ofwca['Name'] == 'MAGRO':
        ExcelApp.row_range_input(data, True)
    else:
        ExcelApp.row_range_input(data)

    """  RATY  """

    for num in range(1, len(r.get('payment'))):
        x_rata = r.get('payment')[num].get('policy_installment_sum_real', '')
        x_ter_platnosci = ExcelApp.date_formatter(r.get('payment', '')[num].get('policy_installment_date_due', ''))
        nr_raty = num + 1
        x_rata_dane = data[:-13] + ['', x_ter_platnosci, x_rata, f_platnosci, ilosc_rat, nr_raty, '', '', '', '',
                                    'api', '', tow_ub]

        if ofwca['Name'] == 'MAGRO':
            ExcelApp.row_range_input(x_rata_dane, True)
        else:
            ExcelApp.row_range_input(x_rata_dane)


end_time = time.time() - start_time
print('\n\n    * Czas zapisania: {:.2f} sekund * '.format(end_time))
time.sleep(30)


"""Opcje zapisania"""

# ExcelApp.DisplayAlerts = False
#
# ## Bez zapisania
# # wb.SaveAs("M:\\Agent baza\\2014 BAZA MAGRO.xlsm")
#
# # Testy
# # wb.SaveAs(path + "\\2014 BAZA MAGRO.xlsm")
#
# """Zamknięcie narazie wyłączone..."""
# # wb.Close()
# ExcelApp.DisplayAlerts = True
#
# # ExcelApp.Application.Quit()
#
# end_time = time.time() - start_time
# print('Czas wykonania: {:.2f} sekund'.format(end_time))